import openpyxl
class ExcelParser:

    def __init__(self, filename):
        self.filename = filename


    def load(self):

        wb = openpyxl.load_workbook(
            self.filename,
            data_only=True
        )

        ws = wb.active

        self.expand_merged_cells(ws)

        return ws


    def expand_merged_cells(self, ws):

        for merged in list(ws.merged_cells.ranges):
            value = ws.cell(merged.min_row, merged.min_col).value
            ws.unmerge_cells(merged.coord)

            for row in range(merged.min_row, merged.max_row + 1):
                for col in range(merged.min_col, merged.max_col + 1):
                    ws.cell(row, col).value = value


    def find_header_row(self, ws):

        for row in range(1, ws.max_row+1):

            values = [
                str(ws.cell(row,c).value)
                for c in range(1,ws.max_column+1)
                if ws.cell(row,c).value
            ]

            text = " ".join(values).lower()


            if "wavelength" in text:
                return row


        raise Exception(
            "No wavelength header found"
        )
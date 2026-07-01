import matplotlib.pyplot as plt
import openpyxl

'''
Requirements:
pip install openpyxl matplotlib


def find_labels(filepath):
    #check row 3 to see whats wavelength, transmission, reflectance, etc.
def plot_transmission(filepath):
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    
    wavelengths = []
    transmissions = []
'''

class SpectrumPoint:
    product: str
    wavelength_nm: float
    metric: str
    value: float


class ExcelReader:

    def load(self, filename):

        wb = openpyxl.load_workbook(
            filename,
            data_only=True
        )

        ws = wb.active

        self.expand_merged_cells(ws)

        return ws


    def expand_merged_cells(self, ws):

        for merged in ws.merged_cells.ranges:

            value = ws.cell(
                merged.min_row,
                merged.min_col
            ).value

            for row in ws[merged.coord]:
                for cell in row:
                    cell.value = value

class HeaderParser:


    def find_header(self, ws):

        for row in range(1, ws.max_row):

            values = [
                str(ws.cell(row,c).value)
                for c in range(1,ws.max_column+1)
                if ws.cell(row,c).value
            ]

            text = " ".join(values).lower()

            if "wavelength" in text:
                return row

        raise Exception(
            "No wavelength column found"
        )
    
class HeaderTree:


    def build(self, ws, header_row):

        headers=[]

        for col in range(1, ws.max_column+1):

            parts=[]

            for row in range(1, header_row+1):

                value = ws.cell(row,col).value

                if value:
                    parts.append(str(value))

            headers.append(
                " | ".join(parts)
            )

        return headers
    
import re


class ColumnClassifier:


    def classify(self, header):

        result = {
            "product": None,
            "metric": None,
            "is_wavelength": False
        }


        if "wavelength" in header.lower():
            result["is_wavelength"]=True


        if "reflectance" in header.lower():
            result["metric"]="Reflectance"


        if "transmission" in header.lower():
            result["metric"]="Transmission"


        products = re.findall(
            r'[A-Z]{2}\d+[A-Z]\d+',
            header
        )

        if products:
            result["product"]=products[0]


        return result
    
import pandas as pd


class Normalizer:


    def convert(
        self,
        ws,
        headers,
        classifications,
        header_row
    ):

        wavelength_col=None

        for i,c in enumerate(classifications):

            if c["is_wavelength"]:
                wavelength_col=i


        rows=[]


        for r in range(
            header_row+1,
            ws.max_row+1
        ):

            wavelength = ws.cell(
                r,
                wavelength_col+1
            ).value


            if wavelength is None:
                continue


            for i,c in enumerate(classifications):

                if c["metric"]:

                    value = ws.cell(
                        r,
                        i+1
                    ).value

                    if value is None:
                        continue


                    rows.append({

                        "product":
                            c["product"],

                        "wavelength_nm":
                            wavelength,

                        "metric":
                            c["metric"],

                        "value":
                            value

                    })


        return pd.DataFrame(rows)
    
    
def dataframe_to_dict(df):

    spectra = {}

    for _, row in df.iterrows():

        product = row["product"]
        metric = row["metric"]
        wavelength = row["wavelength_nm"]
        value = row["value"]


        if product not in spectra:
            spectra[product] = {}


        if metric not in spectra[product]:
            spectra[product][metric] = {}


        spectra[product][metric][wavelength] = value


    return spectra

def dataframe_to_dict(df):

    spectra = {}

    for _, row in df.iterrows():

        product = row["product"]
        metric = row["metric"]
        wavelength = row["wavelength_nm"]
        value = row["value"]


        if product not in spectra:
            spectra[product] = {}


        if metric not in spectra[product]:
            spectra[product][metric] = {}


        spectra[product][metric][wavelength] = value


    return spectra

import json

with open("spectra.json","w") as f:
    json.dump(
        spectra,
        f,
        indent=4
    )

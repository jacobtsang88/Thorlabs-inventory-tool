import tempfile
import unittest
from pathlib import Path

import openpyxl

from excel_parser import ExcelParser


class ExcelParserTests(unittest.TestCase):
    def create_sample_workbook(self, path: Path) -> Path:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "Wavelength"
        ws["B1"] = "Transmission"
        ws["A2"] = 500
        ws["B2"] = 0.82

        ws.merge_cells("C1:D1")
        ws["C1"] = "Merged header"

        wb.save(path)
        return path

    def test_load_expands_merged_cells_and_finds_header_row(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "sample.xlsx"
            self.create_sample_workbook(path)

            parser = ExcelParser(str(path))
            ws = parser.load()

            self.assertEqual(parser.find_header_row(ws), 1)
            self.assertEqual(ws["A2"].value, 500)
            self.assertEqual(ws["B2"].value, 0.82)
            self.assertEqual(ws["C1"].value, "Merged header")
            self.assertEqual(ws["D1"].value, "Merged header")


if __name__ == "__main__":
    unittest.main(verbosity=2)
